"""GamePulse — Flask backend with SQLite. No auth, single shared workspace."""
import json
import os
import sqlite3
from contextlib import contextmanager
from flask import Flask, request, jsonify, send_from_directory, send_file

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
app = Flask(__name__, static_folder=BASE_DIR)
DB_PATH = os.path.join(BASE_DIR, "gamepulse.db")

@contextmanager
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
    finally:
        conn.close()

def init_db():
    with get_db() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS payments (
                id TEXT PRIMARY KEY,
                char_id TEXT, account_id TEXT, char_name TEXT,
                server TEXT, order_id TEXT, channel TEXT,
                amount REAL, region TEXT, tz INTEGER,
                priority TEXT, report_time TEXT, report_local TEXT,
                sla_deadline TEXT, sla_hours REAL,
                desc TEXT, status TEXT,
                created_at TEXT, resolved_at TEXT, notes TEXT
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS feedbacks (
                id TEXT PRIMARY KEY,
                source TEXT, category TEXT,
                char_id TEXT, account_id TEXT, char_name TEXT,
                server TEXT, region TEXT,
                content TEXT, impact TEXT, severity TEXT,
                created_at TEXT, status TEXT
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS liveops (
                id TEXT PRIMARY KEY,
                name TEXT, type TEXT,
                start TEXT, "end" TEXT,
                regions TEXT, note TEXT,
                created_at TEXT
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS networks (
                id TEXT PRIMARY KEY,
                region TEXT, type TEXT, time TEXT,
                affected INTEGER, desc TEXT,
                created_at TEXT
            )
        """)
        conn.commit()

def snake_to_camel(s):
    parts = s.split('_')
    return parts[0] + ''.join(p.capitalize() for p in parts[1:])

def row_to_dict(row):
    return {snake_to_camel(k): row[k] for k in row.keys()}

# ── API ──────────────────────────────────────────────

@app.route("/api/data")
def api_data():
    with get_db() as conn:
        payments = [row_to_dict(r) for r in conn.execute("SELECT * FROM payments ORDER BY created_at DESC").fetchall()]
        feedbacks = [row_to_dict(r) for r in conn.execute("SELECT * FROM feedbacks ORDER BY created_at DESC").fetchall()]
        liveops = [row_to_dict(r) for r in conn.execute("SELECT * FROM liveops ORDER BY created_at DESC").fetchall()]
        networks = [row_to_dict(r) for r in conn.execute("SELECT * FROM networks ORDER BY created_at DESC").fetchall()]
    return jsonify({"payments": payments, "feedbacks": feedbacks, "liveops": liveops, "networks": networks})

@app.route("/api/payments", methods=["POST"])
def api_add_payment():
    p = request.json
    with get_db() as conn:
        conn.execute("""
            INSERT INTO payments (id, char_id, account_id, char_name, server,
                order_id, channel, amount, region, tz, priority,
                report_time, report_local, sla_deadline, sla_hours,
                desc, status, created_at, resolved_at, notes)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (p["id"], p.get("charId",""), p.get("accountId",""), p.get("charName",""),
              p.get("server",""), p.get("orderId",""), p.get("channel",""), p.get("amount",0),
              p.get("region",""), p.get("tz",8), p.get("priority","mid"),
              p.get("reportTime",""), p.get("reportLocal",""), p.get("slaDeadline",""),
              p.get("slaHours",4), p.get("desc",""), p.get("status","pending"),
              p.get("createdAt",""), p.get("resolvedAt"), p.get("notes","")))
        conn.commit()
    return jsonify({"ok": True}), 201

VALID_TRANSITIONS = {
    "pending": ["progress", "deleted"],
    "progress": ["done", "deleted"],
    "done": [],         # final state, no transitions allowed
    "deleted": [],      # final state
    "verified": ["reissued", "deleted"],
    "reissued": ["done", "deleted"],
}

@app.route("/api/payments/<pid>", methods=["PUT"])
def api_update_payment(pid):
    p = request.json
    new_status = p.get("status")
    with get_db() as conn:
        current = conn.execute("SELECT status FROM payments WHERE id=?", (pid,)).fetchone()
        if not current:
            return jsonify({"ok": False, "error": "not found"}), 404
        old_status = current["status"]
        allowed = VALID_TRANSITIONS.get(old_status, [])
        if new_status not in allowed:
            return jsonify({"ok": False, "error": f"Cannot change from '{old_status}' to '{new_status}'"}), 409
        conn.execute("UPDATE payments SET status=?, resolved_at=?, notes=? WHERE id=?",
                     (new_status, p.get("resolvedAt"), p.get("notes",""), pid))
        conn.commit()
    return jsonify({"ok": True})

@app.route("/api/feedbacks", methods=["POST"])
def api_add_feedback():
    f = request.json
    with get_db() as conn:
        conn.execute("""
            INSERT INTO feedbacks (id, source, category, char_id, account_id,
                char_name, server, region, content, impact, severity, created_at, status)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (f["id"], f.get("source",""), f.get("category",""), f.get("charId",""),
              f.get("accountId",""), f.get("charName",""), f.get("server",""),
              f.get("region",""), f.get("content",""), f.get("impact","some"),
              f.get("severity","major"), f.get("createdAt",""), "open"))
        conn.commit()
    return jsonify({"ok": True}), 201

@app.route("/api/liveops", methods=["POST"])
def api_add_liveops():
    lo = request.json
    with get_db() as conn:
        conn.execute("""
            INSERT INTO liveops (id, name, type, start, "end", regions, note, created_at)
            VALUES (?,?,?,?,?,?,?,?)
        """, (lo["id"], lo.get("name",""), lo.get("type",""), lo.get("start",""),
              lo.get("end",""), json.dumps(lo.get("regions",[])), lo.get("note",""),
              lo.get("createdAt","")))
        conn.commit()
    return jsonify({"ok": True}), 201

@app.route("/api/networks", methods=["POST"])
def api_add_network():
    n = request.json
    with get_db() as conn:
        conn.execute("""
            INSERT INTO networks (id, region, type, time, affected, desc, created_at)
            VALUES (?,?,?,?,?,?,?)
        """, (n["id"], n.get("region",""), n.get("type",""), n.get("time",""),
              n.get("affected",0), n.get("desc",""), n.get("createdAt","")))
        conn.commit()
    return jsonify({"ok": True}), 201

# ── Report ────────────────────────────────────────────

@app.route("/api/report")
def api_report():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from datetime import datetime, timedelta

    days = request.args.get("days", "7")
    try:
        days = int(days)
    except ValueError:
        days = 7
    since = (datetime.utcnow() - timedelta(days=days)).isoformat() + "Z"

    with get_db() as conn:
        payments = [row_to_dict(r) for r in conn.execute(
            "SELECT * FROM payments WHERE created_at >= ? AND status != 'deleted' ORDER BY created_at DESC", (since,)
        ).fetchall()]
        feedbacks = [row_to_dict(r) for r in conn.execute(
            "SELECT * FROM feedbacks WHERE created_at >= ? ORDER BY created_at DESC", (since,)
        ).fetchall()]

    wb = Workbook()
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="6C5CE7", end_color="6C5CE7", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def style_header(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

    # ── Sheet 1: Summary ──
    ws1 = wb.active
    ws1.title = "总览"
    total = len(payments)
    resolved = len([p for p in payments if p['status'] == 'done'])
    overdue = len([p for p in payments if p['status'] != 'done' and p.get('slaDeadline') and datetime.fromisoformat(p['slaDeadline'].replace('Z','')) < datetime.utcnow()])
    sla_rate = f"{round((total - overdue) / total * 100, 1)}%" if total > 0 else "N/A"
    avg_min = 0
    done_payments = [p for p in payments if p['status'] == 'done' and p.get('resolvedAt')]
    if done_payments:
        total_min = sum((datetime.fromisoformat(p['resolvedAt'].replace('Z','')) - datetime.fromisoformat(p['createdAt'].replace('Z',''))).total_seconds() / 60 for p in done_payments)
        avg_min = round(total_min / len(done_payments))
    avg_h = f"{avg_min // 60}h {avg_min % 60}m" if avg_min > 0 else "N/A"

    ws1.merge_cells('A1:C1')
    ws1.cell(row=1, column=1, value=f"GamePulse 运营报表（最近 {days} 天）").font = Font(bold=True, size=14)
    ws1.cell(row=2, column=1, value=f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(italic=True, color="888888")

    summary = [
        ("工单总数", total), ("已处理", f"{resolved} ({round(resolved/total*100,1) if total else 0}%)"),
        ("处理中", len([p for p in payments if p['status'] in ('pending','progress')])),
        ("SLA 超时", overdue), ("SLA 达标率", sla_rate), ("平均处理时长", avg_h),
        ("反馈总数", len(feedbacks)),
    ]
    for i, (label, val) in enumerate(summary, 4):
        ws1.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws1.cell(row=i, column=2, value=str(val))
    ws1.column_dimensions['A'].width = 18
    ws1.column_dimensions['B'].width = 24

    # ── Sheet 2: Details ──
    ws2 = wb.create_sheet("工单明细")
    headers2 = ["工单号","角色ID","角色名","区服","订单号","渠道","金额","地区","优先级","状态","问题描述","创建时间","处理耗时","SLA"]
    for c, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=c, value=h)
    style_header(ws2, 1, len(headers2))
    for r, p in enumerate(payments, 2):
        sla = "达标" if p['status'] == 'done' else ("超时" if p.get('slaDeadline') and datetime.fromisoformat(p['slaDeadline'].replace('Z','')) < datetime.utcnow() else "进行中")
        duration = ""
        if p['status'] == 'done' and p.get('resolvedAt'):
            d = (datetime.fromisoformat(p['resolvedAt'].replace('Z','')) - datetime.fromisoformat(p['createdAt'].replace('Z',''))).total_seconds() / 60
            duration = f"{int(d//60)}h {int(d%60)}m"
        row_data = [p['id'], p.get('charId',''), p.get('charName',''), p.get('server',''), p.get('orderId',''),
                    p.get('channel',''), p.get('amount',0), p.get('region',''), p.get('priority',''),
                    {'pending':'待处理','progress':'处理中','done':'已完成','verified':'已核实','reissued':'已补发'}.get(p['status'],p['status']),
                    p.get('desc',''), p['createdAt'][:16] if p.get('createdAt') else '', duration, sla]
        for c, v in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=v).border = thin_border
    for c in range(1, len(headers2)+1):
        ws2.column_dimensions[chr(64+c) if c <= 26 else 'A'].width = 14

    # ── Sheet 3: Analysis ──
    ws3 = wb.create_sheet("分析")
    ws3.cell(row=1, column=1, value="渠道分布").font = header_font
    channels = {}
    for p in payments: channels[p.get('channel','未知')] = channels.get(p.get('channel','未知'), 0) + 1
    for i, (ch, cnt) in enumerate(sorted(channels.items(), key=lambda x: -x[1]), 2):
        ws3.cell(row=i, column=1, value=ch)
        ws3.cell(row=i, column=2, value=cnt)
        ws3.cell(row=i, column=3, value=f"{round(cnt/total*100,1)}%" if total else "0%")

    ws3.cell(row=1, column=5, value="地区分布").font = header_font
    regions = {}
    for p in payments: regions[p.get('region','未知')] = regions.get(p.get('region','未知'), 0) + 1
    for i, (rg, cnt) in enumerate(sorted(regions.items(), key=lambda x: -x[1]), 2):
        ws3.cell(row=i, column=5, value=rg)
        ws3.cell(row=i, column=6, value=cnt)
        ws3.cell(row=i, column=7, value=f"{round(cnt/total*100,1)}%" if total else "0%")

    ws3.cell(row=1, column=9, value="反馈分类").font = header_font
    fbcats = {}
    for f in feedbacks: fbcats[f.get('category','未知')] = fbcats.get(f.get('category','未知'), 0) + 1
    for i, (cat, cnt) in enumerate(sorted(fbcats.items(), key=lambda x: -x[1]), 2):
        ws3.cell(row=i, column=9, value=cat)
        ws3.cell(row=i, column=10, value=cnt)

    for col in [1,5,9]:
        ws3.column_dimensions[chr(64+col)].width = 16

    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"gamepulse_report_{days}d_{datetime.now().strftime('%Y%m%d')}.xlsx")

# ── Static files ─────────────────────────────────────

@app.route("/")
def serve_index():
    return send_from_directory(app.static_folder, "index.html")

@app.route("/<path:path>")
def serve_static(path):
    return send_from_directory(app.static_folder, path)

# Auto-init DB on startup (works in both WSGI and dev modes)
init_db()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5001))
    print(f"🚀 GamePulse running at http://localhost:{port}")
    app.run(host="0.0.0.0", port=port, debug=False)
